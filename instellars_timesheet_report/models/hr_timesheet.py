from datetime import date
from dateutil.relativedelta import relativedelta
from lxml import etree

from odoo import models, fields, api, _
from odoo.addons.web_grid.models.models import END_OF, STEP_BY, START_OF
from odoo.osv import expression
from odoo.tools import date_utils, groupby as groupbyelem
from operator import itemgetter
import itertools
import io
import locale
import base64
import textwrap
from copy import copy
from openpyxl import Workbook
from openpyxl.drawing.image import Image
from openpyxl.drawing.xdr import XDRPositiveSize2D
from openpyxl.utils.units import pixels_to_EMU, cm_to_EMU
from odoo.exceptions import UserError, AccessError,ValidationError,Warning
from openpyxl.drawing.spreadsheet_drawing import OneCellAnchor, AnchorMarker
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font, colors



class ResPartner(models.Model):
    _inherit = 'res.partner'

    sepcific_excel_header = fields.Boolean()


class HrTimesheet(models.Model):

    _inherit = 'account.analytic.line'


    def generate_excel_tm_rport(self):

        if self.env.context.get('grid_anchor'):
            anchor = fields.Date.from_string(self.env.context['grid_anchor'])
        else:
            anchor = date.today() + relativedelta(weeks=-1, days=1, weekday=0)
        span = self.env.context.get('grid_range', 'week')
        validate_to = anchor + END_OF[span]
        validate_from = anchor + START_OF[span] 


        if not self:
            raise UserError(_("There aren't any timesheet to validate"))

        # domain = self.env['ir.rule']._compute_domain(self._name, 'write')  # can write on the timesheet of team employee or all

        employees = self.filtered_domain([]).mapped('employee_id')
        project = self.filtered_domain([]).mapped('project_id')


   

        # tm = self.env['account.analytic.line'].search([('id','in',self.ids)]).filtered(lambda l: l.manager_approval_state == 'approved' and l.validated == False)
        tm = self.env['account.analytic.line'].search([('id','in',self.ids)]).filtered(lambda l: l.manager_approval_state in ['pending_submission','pending_approval','resubmit'] and l.validated == False)

        if tm:
            group_by_employee  = [g for k, g in groupbyelem(tm, itemgetter('employee_id'))]

            emp = [each[0].employee_id.name for each in group_by_employee]
            raise UserError(_("There are some timesheet not approved by manager for employees %s") % (', '.join(filter(None, emp))))


        # employees = employees.sudo().filtered(lambda e: not e.timesheet_validated or e.timesheet_validated < validate_to)

        if not employees:
            raise UserError(_('All selected timesheets for which you are indicated as responsible are already validated.'))


        # analysis_report = ''
        # file_name = ''

        timesheet_repo_lines = []

        for each in employees:
            for l in project:
                analysis_report = ''
                file_name = ''
                # for each in employees:
                # timesheets = self.env['account.analytic.line'].search([('employee_id','=', each.id),('project_id','in',project.ids)],order="date asc").filtered(lambda l: l.manager_approval_state != 'approved' and l.validated == False and l.date <= validate_to and l.date >= validate_from)
                timesheets = self.env['account.analytic.line'].search([('employee_id','=',each.id),('project_id','=',l.id)],order="date asc").filtered(lambda l: l.manager_approval_state == 'approved' and l.validated == False and l.date <= validate_to and l.date >= validate_from)
                if timesheets:
                
                    #Create Workbook and Worksheet
                    wb = Workbook()
                    ws = wb.active
                    ws.title = each.name.replace(' ','_')+'_'+validate_to.strftime('%b_%Y')
                    ws.column_dimensions['A'].width = 22
                    ws.column_dimensions['B'].width = 20
                    ws.column_dimensions['c'].width = 25
                    ws.column_dimensions['D'].width = 25
                    ws.column_dimensions['E'].width = 20
                    ws.column_dimensions['F'].width = 18
                    ws.column_dimensions['G'].width = 30
                    ws.column_dimensions['H'].width = 25
                    ws.column_dimensions['J'].width = 15
                    ws.column_dimensions['K'].width = 15
                    ws.column_dimensions['L'].width = 15
                    # current_row=4
                    #Border


                    border_thin = Border(left=Side(border_style='thin', color='000000'),right=Side(border_style='thin',color='000000'),top=Side(border_style='thin',color='000000'),bottom=Side(border_style='thin',color='000000'))
                    black_color = colors.Color(indexed=0)
                    border_thick = Border(left=Side(border_style='thick', color='000000'),right=Side(border_style='thick',color='000000'),top=Side(border_style='thick',color='000000'),bottom=Side(border_style='thick',color='000000'))
                    ws.merge_cells(start_row=1, start_column=1, end_row=2, end_column=8)

                    if l.partner_id.sepcific_excel_header:

                        #Image placement settings
                        c2e = cm_to_EMU
                        p2e = pixels_to_EMU

                        partnerlogoh = lambda x: c2e((x * 49.77)/99)
                        partnerlogow = lambda x: c2e((x * (18.65-1.71))/10)
                        partnerlogocoloffset = partnerlogow(0.1)
                        partnerlogorowoffset = partnerlogoh(0.1)
                        decoded_img = base64.b64decode(l.partner_id.image_1920)
                        company_logo = Image(io.BytesIO(decoded_img))
                        h, w = 100,250
                        # h, w = company_logo.height, company_logo.width
                        size = XDRPositiveSize2D(p2e(w), p2e(h))
                        marker = AnchorMarker(col=0, colOff=partnerlogocoloffset, row=3, rowOff=partnerlogorowoffset)
                        company_logo.anchor = OneCellAnchor(_from=marker, ext=size)
                        ws.add_image(company_logo)
                        # ws.add_image(product_img,'F'+str(current_row+1))
                        prod_img = ws['A8']
                        prod_img.border = Border(bottom=None,left=None,top=None)
                        ws.merge_cells(start_row=3, start_column=1, end_row=8, end_column=1)

                        comapny = ws.cell(row=1, column=1, value='Timesheets For Month '+ validate_to.strftime('%B %Y'))
                        comapny.alignment = Alignment(horizontal='center',vertical='center')
                        comapny.border = Border(left=Side(border_style=None, color='FF000000'),right=Side(border_style=None,color='FF000000'),top=Side(border_style=None,color='FF000000'),bottom=Side(border_style=None,color='FF000000'))
                        comapny.fill = PatternFill("solid", fgColor="f2dbdb")
                        comapny.font = Font(size=20,name='Calibri',bold=True)

                        pm = ws.cell(row=3, column=3, value='Project Manager:')
                        pm.border = border_thin
                        pm.alignment = Alignment(horizontal='center',vertical='center')
                        pm.font = Font(size=11,name='Arial',bold=True)
                        ws.merge_cells(start_row=3, start_column=3, end_row=3, end_column=5)

                        pm_val = ws.cell(row=3, column=6, value='')
                        pm_val.alignment = Alignment(horizontal='center',vertical='center')
                        pm_val.fill = PatternFill("solid", fgColor="ccffcc")
                        pm_val.font = Font(size=11,name='Arial',bold=True)
                        ws.merge_cells(start_row=3, start_column=6, end_row=3, end_column=8)
                        pm_val.border = border_thin

                        agency_name = ws.cell(row=4, column=3, value='Agency Name:')
                        agency_name.alignment = Alignment(horizontal='center',vertical='center')
                        agency_name.font = Font(size=11,name='Arial',bold=True)
                        ws.merge_cells(start_row=4, start_column=3, end_row=4, end_column=5)
                        agency_name.border = border_thin

                        agency_name_val = ws.cell(row=4, column=6, value=self.env.user.company_id.name)
                        agency_name_val.alignment = Alignment(horizontal='center',vertical='center')
                        agency_name_val.fill = PatternFill("solid", fgColor="ccffcc")
                        agency_name_val.font = Font(size=11,name='Arial',bold=True)
                        ws.merge_cells(start_row=4, start_column=6, end_row=4, end_column=8)
                        agency_name_val.border = border_thin

                        service_order_no = ws.cell(row=5, column=3, value='Service Order No(Optional Field):')
                        service_order_no.alignment = Alignment(horizontal='center',vertical='center')
                        service_order_no.font = Font(size=11,name='Arial',bold=True)
                        ws.merge_cells(start_row=5, start_column=3, end_row=5, end_column=5)
                        service_order_no.border = border_thin

                        service_order_no_val = ws.cell(row=5, column=6, value='')
                        service_order_no_val.alignment = Alignment(horizontal='center',vertical='center')
                        service_order_no_val.fill = PatternFill("solid", fgColor="ccffcc")
                        service_order_no_val.font = Font(size=11,name='Arial',bold=True)
                        ws.merge_cells(start_row=5, start_column=6, end_row=5, end_column=8)
                        service_order_no_val.border = border_thin

                        month_of_work_h = ws.cell(row=7, column=3, value='Month of work')
                        month_of_work_h.alignment = Alignment(horizontal='center',vertical='center')
                        month_of_work_h.font = Font(size=11,name='Arial',bold=True)
                        ws.merge_cells(start_row=7, start_column=3, end_row=7, end_column=4)
                        month_of_work_h.border = border_thin

                        month_of_work_value = ws.cell(row=8, column=3, value=validate_to.strftime('%B'))
                        month_of_work_value.alignment = Alignment(horizontal='center',vertical='center')
                        # month_of_work_value.fill = PatternFill("solid", fgColor="ccffcc")
                        month_of_work_value.font = Font(size=11,name='Arial',bold=True)
                        ws.merge_cells(start_row=8, start_column=3, end_row=8, end_column=4)
                        month_of_work_value.border = border_thin

                        year_h = ws.cell(row=7, column=5, value='Year')
                        year_h.alignment = Alignment(horizontal='center',vertical='center')
                        year_h.font = Font(size=11,name='Arial',bold=True)
                        ws.merge_cells(start_row=7, start_column=5, end_row=7, end_column=5)
                        year_h.border = border_thin

                        year_value = ws.cell(row=8, column=5, value=validate_to.strftime('%Y'))
                        year_value.alignment = Alignment(horizontal='center',vertical='center')
                        year_value.fill = PatternFill("solid", fgColor="ccffcc")
                        year_value.font = Font(size=11,name='Arial',bold=True)
                        ws.merge_cells(start_row=8, start_column=5, end_row=8, end_column=5)
                        year_value.border = border_thin

                        sub_con_name = ws.cell(row=7, column=6, value='Sub Con Name')
                        sub_con_name.alignment = Alignment(horizontal='center',vertical='center')
                        sub_con_name.font = Font(size=11,name='Arial',bold=True)
                        ws.merge_cells(start_row=7, start_column=6, end_row=7, end_column=8)
                        sub_con_name.border = border_thin

                        sub_con_name = ws.cell(row=8, column=6, value=each.name)
                        sub_con_name.alignment = Alignment(horizontal='center',vertical='center')
                        sub_con_name.fill = PatternFill("solid", fgColor="ccffcc")
                        sub_con_name.font = Font(size=11,name='Arial',bold=True)
                        ws.merge_cells(start_row=8, start_column=6, end_row=8, end_column=8)
                        sub_con_name.border = border_thin


                    else:
                        ##consultant details
                        consultant_det_h = ws.cell(row=3, column=1, value='Consultant Details')
                        consultant_det_h.alignment = Alignment(horizontal='center',vertical='center')
                        consultant_det_h.border = border_thin
                        consultant_det_h.fill = PatternFill("solid", fgColor="a5b6ca")
                        consultant_det_h.font = Font(size=11,name='Arial',bold=True)
                        ws.merge_cells(start_row=3, start_column=1, end_row=3, end_column=2)

                        consultant_details = ws.cell(row=4, column=1, value='1.Consultant Name: '+timesheets[0].employee_id.name+'\n2.Skill: '+timesheets[0].employee_id.employee_skill_ids[0].skill_type_id.name if timesheets[0].employee_id.employee_skill_ids else '' +'\n3.Deployement Date:\n4.Timesheet Submitted Date:')
                        consultant_details.alignment = Alignment(horizontal='left',vertical='center')
                        consultant_details.border = border_thin
                        consultant_details.font = Font(size=11,name='Arial')
                        ws.merge_cells(start_row=4, start_column=1, end_row=7, end_column=2)

                        #client details

                        client_det_h = ws.cell(row=3, column=4, value='Client Details')
                        client_det_h.alignment = Alignment(horizontal='center',vertical='center')
                        client_det_h.border = border_thin
                        client_det_h.fill = PatternFill("solid", fgColor="a5b6ca")
                        client_det_h.font = Font(size=11,name='Arial',bold=True)
                        ws.merge_cells(start_row=3, start_column=4, end_row=3, end_column=5)

                        client_details = ws.cell(row=4, column=4, value='1.Client Name: '+l.partner_id.name if l.partner_id else '' +'\n2.Client Project Manager:'+l.partner_id.name if l.partner_id else ''+'\n3.Project Name: '+l.name if l.name else ''+'\n4.Timesheet Submitted Date:')
                        client_details.alignment = Alignment(horizontal='left',vertical='center')
                        client_details.border = border_thin
                        client_details.font = Font(size=11,name='Arial')
                        ws.merge_cells(start_row=4, start_column=4, end_row=7, end_column=5)


                        #Vendor Details

                        vendor_det_h = ws.cell(row=3, column=7, value='Vendor Details')
                        vendor_det_h.alignment = Alignment(horizontal='center',vertical='center')
                        vendor_det_h.border = border_thin
                        vendor_det_h.fill = PatternFill("solid", fgColor="a5b6ca")
                        vendor_det_h.font = Font(size=11,name='Arial',bold=True)
                        ws.merge_cells(start_row=3, start_column=7, end_row=3, end_column=8)

                        vendor_details = ws.cell(row=4, column=7, value='1.Vendor Name:'+self.env.company.name+'\n2.Vendor Account Manager:\n3.Vendor Project Co-ordinator :\n4.Timesheet Verified Date:')
                        vendor_details.alignment = Alignment(horizontal='left',vertical='center')
                        vendor_details.border = border_thin
                        vendor_details.font = Font(size=11,name='Arial')
                        ws.merge_cells(start_row=4, start_column=7, end_row=7, end_column=8)

                        comapny = ws.cell(row=1, column=1, value='Timesheets For Month '+ validate_to.strftime('%B %Y'))
                        comapny.alignment = Alignment(horizontal='center',vertical='center')
                        comapny.border = Border(left=Side(border_style=None, color='FF000000'),right=Side(border_style=None,color='FF000000'),top=Side(border_style=None,color='FF000000'),bottom=Side(border_style=None,color='FF000000'))
                        comapny.fill = PatternFill("solid", fgColor="f2dbdb")
                        comapny.font = Font(size=20,name='Calibri',bold=True)



                    # Defining the Table Caolumn headings
                    date_of_work = ws.cell(row=11, column=1, value="Date Of Work")
                    date_of_work.alignment = Alignment(horizontal='center',vertical='center')
                    date_of_work.font = Font(size=11,name='Arial',bold=True)
                    date_of_work.border = border_thin

                    day = ws.cell(row=11, column=2, value="Day")
                    day.alignment = Alignment(horizontal='center',vertical='center')
                    day.font = Font(size=11,name='Arial',bold=True)
                    day.border = border_thin

                    status = ws.cell(row=11, column=3, value="Status")
                    status.alignment = Alignment(horizontal='center',vertical='center')
                    status.font = Font(size=11,name='Arial',bold=True)
                    status.border = border_thin

                    duration_days = ws.cell(row=11, column=4, value="No Of Days")
                    duration_days.alignment = Alignment(horizontal='center',vertical='center')
                    duration_days.font = Font(size=11,name='Arial',bold=True)
                    duration_days.border = border_thin

                    # region = ws.cell(row=11, column=4, value="Region")
                    # region.alignment = Alignment(horizontal='center',vertical='center')
                    no_of_days = ws.cell(row=11, column=5, value="Delivery Site")
                    no_of_days.alignment = Alignment(horizontal='center',vertical='center')
                    no_of_days.font = Font(size=11,name='Arial',bold=True)
                    no_of_days.border = border_thin

                    duration = ws.cell(row=11, column=6, value="Duration")
                    duration.alignment = Alignment(horizontal='center',vertical='center')
                    duration.font = Font(size=11,name='Arial',bold=True)
                    duration.border = border_thin

                    description = ws.cell(row=11, column=7, value="Description")
                    description.alignment = Alignment(horizontal='center',vertical='center')
                    description.font = Font(size=11,name='Arial',bold=True)
                    description.border = border_thin

                    billable_hrs = ws.cell(row=11, column=8, value="Billable Hours")
                    billable_hrs.alignment = Alignment(horizontal='center',vertical='center')
                    billable_hrs.font = Font(size=11,name='Arial',bold=True)
                    billable_hrs.border = border_thin



                    row_c=12

                    for each_line in timesheets:
                        if each_line.status == 'full_day':
                            sts = 'Full Day'
                        elif each_line.status == 'half_day':
                            sts = 'Half Day'
                        elif each_line.status == 'absent':
                            sts = 'Absent'
                        elif each_line.status == 'weekend':
                            sts = 'Weekend'
                        elif each_line.status == 'public_holiday':
                            sts = 'Public HoliDay'
                        elif each_line.status == 'comp_off':
                            sts = 'Comp Off'
                        elif each_line.status == 'business_travel':
                            sts = 'Business Travel'
                        else:
                            sts = ''

                        date_of_work = ws.cell(row=row_c, column=1, value=each_line.date.strftime("%B %d, %Y"))
                        date_of_work.alignment = Alignment(horizontal='center',vertical='center')
                        date_of_work.border = border_thin

                        day = ws.cell(row=row_c, column=2, value=each_line.day)
                        day.alignment = Alignment(horizontal='center',vertical='center')
                        day.border = border_thin

                        status = ws.cell(row=row_c, column=3, value=sts)
                        status.alignment = Alignment(horizontal='center',vertical='center')
                        status.border = border_thin

                        duration_days = ws.cell(row=row_c, column=4, value="{0:.1f}".format(each_line.duration_days))
                        duration_days.alignment = Alignment(horizontal='center',vertical='center')
                        duration_days.border = border_thin

                        no_of_days = ws.cell(row=row_c, column=5, value=each_line.delivery_site.name)
                        no_of_days.alignment = Alignment(horizontal='center',vertical='center')
                        no_of_days.border = border_thin

                        duration = ws.cell(row=row_c, column=6, value=each_line.unit_amount)
                        duration.alignment = Alignment(horizontal='center',vertical='center')
                        duration.border = border_thin

                        description = ws.cell(row=row_c, column=7, value=each_line.name)
                        description.alignment = Alignment(horizontal='center',vertical='center')
                        description.border = border_thin

                        billable_hrs = ws.cell(row=row_c, column=8, value='')
                        billable_hrs.alignment = Alignment(horizontal='center',vertical='center')
                        billable_hrs.border = border_thin

                        row_c +=1


                    total_duraton = sum([rec.unit_amount for rec in timesheets])

                    tot_duration = ws.cell(row=row_c, column=6, value=total_duraton)
                    tot_duration.alignment = Alignment(horizontal='center',vertical='center')
                    tot_duration.font = Font(size=11,name='Arial',bold=True,underline='single')
                    tot_duration.border = border_thin


                    footer_subcon = ws.cell(row=row_c+2, column=1, value='Sub Con\'s Signature:')
                    footer_subcon.alignment = Alignment(horizontal='left',vertical='top')
                    footer_subcon.border = border_thin
                    footer_subcon.font = Font(size=11,name='Arial',bold=True)
                    ws.merge_cells(start_row=row_c+2, start_column=1, end_row=row_c+8, end_column=2)

                    footer_pm = ws.cell(row=row_c+2, column=6, value='Project Manager\'s Signature for approval:')
                    footer_pm.alignment = Alignment(horizontal='left',vertical='top')
                    footer_pm.border = border_thin
                    footer_pm.font = Font(size=11,name='Arial',bold=True)
                    ws.merge_cells(start_row=row_c+2, start_column=6, end_row=row_c+8, end_column=8)


                    fp = io.BytesIO()
                    wb.save(fp)
                    excel_file = base64.encodestring(fp.getvalue())
                    analysis_report = excel_file
                    # self.margin_analysis_printed = True
                    file_name = each.name.replace(' ','_')+'_'+validate_to.strftime('%b_%Y')
                    fp.close()

                    timesheet_repo_lines.append({'employee_id':timesheets[0].employee_id.id,'project_id':l.id,'timesheet_report':analysis_report,'file_name':file_name})


        validation = self.env['timesheet.report.wiz'].create({
            'from_date': validate_from,
            'to_date': validate_to,
            'timesheet_report_line_ids': [
                (0, 0, {'employee_id': rec['employee_id'],'project_id':rec['project_id'],'timesheet_report':rec['timesheet_report'],'file_name':rec['file_name']}) for rec in timesheet_repo_lines
            ]
            
        })
        # validation = self.env['timesheet.validation'].create({
        #     'validation_date': validate_to,
        #     'validation_line_ids': [
        #         (0, 0, {'employee_id': employee.id}) for employee in employees
        #     ]
        # })


        # if analysis_report:
        #     validation.write({'file_name':file_name,'timesheet_report':analysis_report})


        return {
            'name': _('Timesheet Report'),
            'type': 'ir.actions.act_window',
            'target': 'new',
            'res_model': 'timesheet.report.wiz',
            'res_id': validation.id,
            'views': [(False, 'form')],
        }


